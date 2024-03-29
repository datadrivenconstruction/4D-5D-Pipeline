# -*- coding: utf-8 -*-
"""Copy of DataDrivenConstruction Pipeline 5D QTO.ipynb

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/16056Jf-ivOOsjD0PuGbOTIwDp4ve1n2X
"""

###
# Pipeline:  5D QTO
# URI: https://DataDrivenConstruction.io/
# Description: Grouping the model by parameters and filling the table to create 5D data
# DataDrivenConstruction
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
###

import time, json, os, re, numpy as np, subprocess, warnings, pandas as pd
warnings.simplefilter(action='ignore', category=FutureWarning)
import xml.etree.ElementTree as ET
from slugify import slugify
import requests
from openpyxl import load_workbook

start_time = time.time()

# Folders where the conversion files are located
path = 'C:\\DataDrivenConstruction\\Input\\'
outpath = path + 'Output\\'

pathn = path + '/OpenCostEstimate'
try:
    os.mkdir(pathn)
except:
    pass
# Properties for which we want to collect data on the amount of volume
propstr =  ['Area', 'Volume', 'Width', 'Length', 'Perimeter', 'öööasdöööfake']
search_parameters = ['Type Name', 'ObjectType', 'Reference', 'Familie und Typ', 'Familie']

# Main function for grouping data and saving a file
def crtable(filename):
        filenamep = outpath + filename
        df = pd.read_csv(filenamep, low_memory=False)
        filedae = outpath + filename[:-8]+'dae'
        print(filedae)
        #    Fetching only numbers from string values of volumetric parameters
        
        propindf, sp = [], []   
        
        #grouping by element types for different formats
        for el in search_parameters:
            if el in df.columns:
                sp.append(el)
        search_parameter = sp[0]

        # Converting all "propstr" values in columns to numeric values
        for el in propstr:
            if el in df.columns:
                propindf.append(el)
        def find_number(text):
            num = re.findall(r'[0-9]+', text)
            return ".".join(num)
        for el in propindf:
            df[el] = df[el].astype(str)
            df[el] = df[el].apply(lambda x: find_number(x))
            df[el] = pd.to_numeric(df[el], errors='coerce')
            df[el] = df[el].replace(np.nan, 0)
            df[el] = df[el].replace('None', 0)
            df[el] = df[el].fillna(0)
        try:
                df[el] = df[el].astype(float)
        except:
                pass

        # Summation of all data that are grouped by search_parameter located in the propindf columns
        df1=pd.pivot_table(df, index=[search_parameter],values=propindf,aggfunc=np.sum)
        df1 = df1.add_prefix('Sum of ')

        # Determination of the number of elements in groups
        df2= df.groupby([search_parameter])[propindf[0]].agg(['count'])
        dfallpar = pd.DataFrame()    
        df['Unnamed: 0'] = df['Unnamed: 0'].astype(str)
        comma = lambda x: ', '.join(x.unique())
        df3 = df.groupby([search_parameter]) .agg({'Unnamed: 0': comma})
        
        # Collecting data into one dataframe
        dfallpar = pd.concat([df2, df1, df3], axis=1)
        dfallpar.rename(columns=({ 'Unnamed: 0': 'Id´s', 'count': 'Amount'}), inplace=True,)
     
        # Use and download a sample excel file
        url = 'https://github.com/DataDrivenConstruction/Open-Estimation/raw/main/OpenEstimator.xlsx'
        r = requests.get(url)
        excelf = pathn + '/' + 'OCE_' + filename+'.xlsx'         
        with open(excelf, 'wb') as f:
                f.write(r.content)

        # Saving data to file
        book = load_workbook(excelf)
        writer = pd.ExcelWriter(excelf, engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        dfallpar.to_excel(writer, 'BIM_Data')
        writer.save()

        dfallpar['type'] = dfallpar.index
        dfallpar.insert(0, 'Group Key', dfallpar['type'].apply(slugify))
        print("File created: " + excelf)
        
        # Start sorting geometry from DAE file
        # Formation of a data tree from the DAE format
        daegrpath = pathn + '/' + 'DAEgroups_' + filename[:-9] 
        try:
            os.mkdir(daegrpath)
        except:
            pass

        # If the ID of an element from the group_ids_str list that was found earlier matches,
        # all elements with this ID are found in the DAE file, and all other elements are deleted
        filedaearr = []

        for index, row in dfallpar.iterrows():
            fileObject = open(filedae, "r")
            treeq = ET.parse(fileObject)
            root = treeq.getroot()
            ET.register_namespace("", "http://www.collada.org/2005/11/COLLADASchema")
            geom_list = []
            group_ids_str = []
            group_ids_str = re.findall(r'\d+', row['Id´s'])
            for node in root.findall('.//{http://www.collada.org/2005/11/COLLADASchema}node'):
                    tree = treeq
                    if node.attrib['id'] in group_ids_str:
                        try:
                            url = list(node)[0].get('url')
                            geom_list.append(url[1:])
                        except:
                            pass
                    else:
                            try:
                                    nd = node.find(
                                            '{http://www.collada.org/2005/11/COLLADASchema}instance_geometry')
                                    node.remove(nd)
                            except:
                                    0
            for geomet in root.findall('.//{http://www.collada.org/2005/11/COLLADASchema}geometry'):
                        if geomet.attrib['id'] in geom_list:
                                0
                        else:
                                try:
                                    md = geomet.find(
                                            '{http://www.collada.org/2005/11/COLLADASchema}mesh')
                                    geomet.remove(md)
                                except:
                                    pass

            # Formation of a new name for the DAE file with grouped elements
            #words_pattern = '[a-zA-Z10-9]+'
            invalid = '<>:"/\|?* '
            for char in invalid:
                index = index.replace(char, '')
            regw = index + '.dae'
            filedaena = daegrpath + '/' + regw
            with open(filedaena, 'w') as f:
                    tree.write(f, encoding='unicode')
            #filedaearr.append("""=HYPERLINK("["""+"/" + "DAEgroups_" + filename[:-9] + "/" + regw + "]" + regw +"""")""")
            filedaearr.append('=HYPERLINK(LEFT(CELL("filename",A1),FIND("[",CELL("filename",A1))-1)&"' + "DAEgroups_" + filename[:-9] + '\\' + regw +'","'+ regw + '")')
        dfallpar.drop(columns=['type'])
        dfallpar.insert(7, "Group geometry in DAE, file hyperlink *.dae", filedaearr)
        with open(excelf, 'wb') as f:
                f.write(r.content)

        # Saving data to file
        book = load_workbook(excelf)
        writer = pd.ExcelWriter(excelf, engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        dfallpar.to_excel(writer, 'BIM Data')
        writer.save()

# Function execution cycle for all CSV files in the folder
for filename in os.listdir(outpath):
    if filename.endswith("csv"): 
        try:    
            crtable(filename)
        except:
            pass
            
print("--- %s seconds ---" % (time.time() - start_time))
# Saving data to a ZIP file for downloading to a computer
#!zip -r /content/file.zip /content/rvt